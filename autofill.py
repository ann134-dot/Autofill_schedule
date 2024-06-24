# 1. get data from a schedule file
# data[name] = 
# class Person -> role, name, schedule
# schedule: days = [dict[date] = color] 
# theme=7 - day, theme=4 - night, 0 - none, tint=-0.3499862666707358 gray - vacation 1
# rgb='FFFF0000 - holidays, theme=5 - weekend, theme=9 - 8h, rgb='FF6699FF' - layoff 2
# 2. parse it to the table file
# DOES NOT WORK FOR FEBRUARY

import openpyxl
from openpyxl.styles import PatternFill
from person import Person
import datetime

people = []
path = 'График апрель 2024 new.xlsx'
path_wr_table = 'Табель май 2024г. АО.xlsx'
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active


for i in range(17, 24): 
    schedule = {}
    name = sheet_obj.cell(row = i, column = 2).value
    role = sheet_obj.cell(row = i, column = 3).value
    if 'L1' in role:
        for j in range(4, 20):
            color1 = sheet_obj.cell(row = i, column = j).fill.start_color
            date1 = sheet_obj.cell(row = 10, column = j).value
            # for the second half of the month 
            color2 = sheet_obj.cell(row = i+17, column = j).fill.start_color
            date2 = sheet_obj.cell(row = 27, column = j).value
            if date2 is None or date1 is None:
                continue
            
            # print(color1 )
            # day off
            if color1.rgb == 'FFFFFFFF' and color1.type=='rgb':
                schedule[date1] = '0'
            # day
            elif color1.rgb == 'FFFFE699' and color1.type=='rgb':
                schedule[date1] = '7'
            # night
            elif color1.rgb == 'FF2F5597' and color1.type=='rgb':
                schedule[date1] = '4'
            # vacation
            elif color1.rgb == 'FFA6A6A6' and color1.type=='rgb':
                schedule[date1] = '1'
            # layoff 
            elif color1.rgb == 'FF6699FF' and color1.type=='rgb':
                schedule[date1] = '2'

            
            # day off
            if color2.rgb == 'FFFFFFFF' and color2.type=='rgb':
                schedule[date2] = '0'
            # day
            elif color2.rgb == 'FFFFE699' and color2.type=='rgb':
                schedule[date2] = '7'
            # night
            elif color2.rgb == 'FF2F5597' and color2.type=='rgb':
                schedule[date2] = '4'
            # vacation
            elif color2.rgb == 'FFA6A6A6' and color2.type=='rgb':
                schedule[date2] = '1'
            # layoff 
            elif color2.rgb == 'FF6699FF' and color2.type=='rgb':
                schedule[date2] = '2'

    else:
        schedule = None   

    person = Person(name, role, schedule)
    people.append(person)
    # print('days: ', person.count_day())
    # print('nights: ', person.count_night())
    # print('total: ', person.count_total_work())
    
    # print('vacation: ', person.count_vacation())
    # for key, val in person.schedule.items():
    #     print(key, val)


# exit()
wb_obj_wr = openpyxl.load_workbook(path_wr_table)
sheet_obj_wr = wb_obj_wr.active  

start_row = 26
end_row = 40
index1 = 4
index2 = 4
is_night_shift1 = False
is_night_shift2 = False

for person in people:
    i = start_row
    print('name: ',person.name,' i: ', i)
    sheet_obj_wr.cell(row=i, column=3).value = person.name
    sheet_obj_wr.cell(row=i, column=2).value = person.role
    day = person.count_day()
    sheet_obj_wr.cell(row=i, column=36).value = day[0]
    sheet_obj_wr.cell(row=i+1, column=36).value = day[1]
    vac = person.count_vacation() 
    if vac:
        sheet_obj_wr.cell(row=i, column=37).value = vac
    night = person.count_night()
    sheet_obj_wr.cell(row=i, column=43).value = night[0]
    sheet_obj_wr.cell(row=i+1, column=43).value = night[1]
    sheet_obj_wr.cell(row=i, column=44).value = person.count_total_work()[0]
    sheet_obj_wr.cell(row=i+1, column=44).value = person.count_total_work()[1]

    for key, val in person.schedule.items():
        day = key.day
        print('day:value =',day,val)

        if day <= 15:
            if is_night_shift1:
                is_night_shift1 = False
                continue
            cell = sheet_obj_wr.cell(row=i, column=index1)
            row = i
            print('index1 ', index1)
            index = index1
            if int(val) != 4:
                index1 +=2
            else:
                index1 +=4
                is_night_shift1 = True
        else:
            if is_night_shift2:
                is_night_shift2 = False
                continue
            cell = sheet_obj_wr.cell(row=i+1, column=index2)
            row=i+1
            print('index2 ', index2)
            index = index2
            if int(val) != 4:
                index2 +=2
            else:
                index2 +=4
                is_night_shift2 = True
    
            
        if int(val) == 7:
            sheet_obj_wr.merge_cells(start_row=row, start_column=index, end_row=row, end_column=index+1)
            cell.value = 11
            cell.fill = PatternFill(start_color="00FFFF00", fill_type="solid")
    
        elif int(val) == 4:
            cell.value = 1
            cell.fill = PatternFill(start_color="fff2cc", fill_type="solid")

            sheet_obj_wr.cell(row=row, column=index+1).value = 2
            sheet_obj_wr.cell(row=row, column=index+1).fill = PatternFill(start_color="b4c7e7", fill_type="solid")
            
            if key.day >=30 or key.day == 15:
                print('shifted night hours')
                continue

            sheet_obj_wr.cell(row=row, column=index+2).value = 5
            sheet_obj_wr.cell(row=row, column=index+2).fill = PatternFill(start_color="b4c7e7", fill_type="solid")
            
            sheet_obj_wr.cell(row=row, column=index+3).value = 3
            sheet_obj_wr.cell(row=row, column=index+3).fill = PatternFill(start_color="fff2cc", fill_type="solid")
        elif int(val) == 0:
            if index == 4: 
                if int(person.schedule[key+datetime.timedelta(days=1)]) == 0:
                    print('shifted 3-5hours to be added...')
                    if key.day == 15:
                        row +=1
                    sheet_obj_wr.cell(row=row, column=4).value = 5
                    sheet_obj_wr.cell(row=row, column=4).fill = PatternFill(start_color="b4c7e7", fill_type="solid")
                    # print(row,sheet_obj_wr.cell(row=row, column=5).value  )
                    sheet_obj_wr.cell(row=row, column=5).value = 3
                    sheet_obj_wr.cell(row=row, column=5).fill = PatternFill(start_color="fff2cc", fill_type="solid")                    
                    continue

            sheet_obj_wr.merge_cells(start_row=row, start_column=index, end_row=row, end_column=index+1)
            cell.value = None
            cell.fill = PatternFill(start_color="ffffff", fill_type="solid")
        # vacation
        elif int(val) == 1:
            sheet_obj_wr.merge_cells(start_row=row, start_column=index, end_row=row, end_column=index+1)
            cell.value = 'O'
            cell.fill = PatternFill(start_color="3399ff", fill_type="solid")  
        #layoff
        elif int(val) == 2:
            sheet_obj_wr.merge_cells(start_row=row, start_column=index, end_row=row, end_column=index+1)
            cell.value = 'У'
            cell.fill = PatternFill(start_color="9999ff", fill_type="solid")  
        # break

    index1 = 4
    index2 = 4
    is_night_shift1 = False
    is_night_shift2 = False
    start_row+=2
    if start_row >= end_row:
        print('end row: ', start_row,' name: ', person.name)
    


wb_obj_wr.save(path_wr_table)
    
    




